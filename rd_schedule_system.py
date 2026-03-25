#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════╗
║    POST OFFICE RD SCHEDULE AUTOMATION SYSTEM                ║
║    ─────────────────────────────────────────                ║
║    Helps RD agents track payments, generate deposit lists,  ║
║    and export schedules for 1000+ accounts efficiently.     ║
║                                                             ║
║    Author : Senior System Architect                         ║
║    Version: 1.0.0                                           ║
║    Python : 3.8+                                            ║
╚══════════════════════════════════════════════════════════════╝

ARCHITECTURE OVERVIEW:
    ┌─────────────┐     ┌──────────────┐     ┌─────────────┐
    │  Excel File  │────▶│  DataLoader  │────▶│  AccountMgr │
    │  (Input)     │     │  (pandas)    │     │  (Business) │
    └─────────────┘     └──────────────┘     └──────┬──────┘
                                                     │
                         ┌──────────────┐            │
                         │ DepositList  │◀───────────┘
                         │ Generator    │
                         └──────┬───────┘
                                │
                    ┌───────────┴───────────┐
                    │  ExcelExporter        │
                    │  (openpyxl formatted) │
                    └───────────────────────┘
"""

import os
import re
import sys
from datetime import datetime, date
from dataclasses import dataclass, field, replace
from typing import List, Optional, Tuple, Dict
from enum import Enum
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from tkinter.scrolledtext import ScrolledText

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════
# SECTION 1: DOMAIN MODELS
# ═══════════════════════════════════════════════════════════════

class Priority(Enum):
    """Payment priority based on account opening date.

    Business Rule:
        - Account opened on/before 15th → HIGH (must pay by 15th)
        - Account opened after 15th → NORMAL (pay anytime in month)
    """
    HIGH = "HIGH"
    NORMAL = "NORMAL"


class AccountStatus(Enum):
    """Lifecycle status of an RD account."""
    ACTIVE = "ACTIVE"
    MATURED = "MATURED"       # 60 installments completed (can still be extended)
    DEFAULTED = "DEFAULTED"   # Missed payment(s)
    EXPIRED = "EXPIRED"       # Due > 6 months overdue — cannot be continued


@dataclass
class RDAccount:
    """Represents a single Recurring Deposit account.

    This is the core domain entity. Each account tracks one customer's
    RD with the post office, including payment history and status.

    Business Rules:
        - Accounts with 60+ months are MATURED but can be extended.
        - Accounts overdue by more than 6 months are EXPIRED (cannot continue).

    Attributes:
        account_no: Unique account identifier
        name: Name of the depositor
        denomination: Fixed monthly deposit amount (₹)
        months_paid: Number of installments completed (can exceed 60 if extended)
        due_date: Next RD installment due date
        priority: HIGH if opened ≤15th, NORMAL otherwise
        status: ACTIVE / MATURED / DEFAULTED / EXPIRED
        is_paid_this_month: Whether payment is marked for current cycle
        has_default: Whether account has any missed payments
    """
    account_no: str
    name: str
    denomination: float
    months_paid: int
    due_date: Optional[date]
    priority: Priority = Priority.NORMAL
    status: AccountStatus = AccountStatus.ACTIVE
    is_paid_this_month: bool = False
    has_default: bool = False

    def __post_init__(self):
        """Auto-calculate priority and status after initialization."""
        from dateutil.relativedelta import relativedelta

        # Determine priority from due date
        if self.due_date and self.due_date.day <= 15:
            self.priority = Priority.HIGH

        today = date.today()

        # Check expiry first: due > 6 months overdue → EXPIRED (cannot continue)
        if self.due_date and self.due_date < today:
            six_months_ago = today - relativedelta(months=6)
            if self.due_date <= six_months_ago:
                self.has_default = True
                self.status = AccountStatus.EXPIRED
                return

        # Check maturity (60 installments = 5 years, but can be extended)
        if self.months_paid >= 60:
            self.status = AccountStatus.MATURED

        # Check for overdue/default (matured accounts can also be defaulted)
        if self.due_date and self.due_date < today:
            self.has_default = True
            if self.status != AccountStatus.MATURED:
                self.status = AccountStatus.DEFAULTED

    @property
    def remaining_months(self) -> int:
        """Months remaining until initial maturity (60). Returns 0 if extended."""
        return max(0, 60 - self.months_paid)

    @property
    def is_matured(self) -> bool:
        """True if 60+ installments completed (may still be active/extended)."""
        return self.months_paid >= 60

    @property
    def is_expired(self) -> bool:
        """True if due date is more than 6 months overdue — cannot be continued."""
        return self.status == AccountStatus.EXPIRED

    @property
    def is_overdue(self) -> bool:
        if not self.due_date:
            return False
        return self.due_date < date.today() and not self.is_expired


@dataclass
class DepositList:
    """A batch of accounts grouped for a single post office deposit.

    Business Rule:
        - Total must not exceed ₹20,000 (configurable)
        - Prefer round totals: ₹20K, ₹19K, ₹18K, ₹10K
        - High priority accounts are selected first
    """
    accounts: List[RDAccount] = field(default_factory=list)
    max_amount: float = 20000

    @property
    def total(self) -> float:
        return sum(a.denomination for a in self.accounts)

    @property
    def remaining_capacity(self) -> float:
        return self.max_amount - self.total

    @property
    def count(self) -> int:
        return len(self.accounts)

    def can_add(self, account: RDAccount) -> bool:
        return account.denomination <= self.remaining_capacity

    def add(self, account: RDAccount) -> bool:
        if self.can_add(account):
            self.accounts.append(account)
            return True
        return False


# ═══════════════════════════════════════════════════════════════
# SECTION 2: DATA LOADER (Excel → Domain Models)
# ═══════════════════════════════════════════════════════════════

class DataLoader:
    """Reads and validates customer data from Excel files.

    Handles real-world messy data:
        - Empty rows, NaN values
        - Denomination format: "2,000.00 Cr." → 2000.0
        - Date parsing with multiple formats
        - Duplicate account detection
    """

    # Expected column mappings (flexible matching)
    COLUMN_MAP = {
        "account_no": ["account no", "account_no", "accountno", "acc no", "acno"],
        "name": ["account name", "name", "depositor", "customer name", "account_name"],
        "denomination": ["denomination", "amount", "deposit", "monthly amount"],
        "months_paid": ["month paid upto", "months paid", "months_paid", "installments"],
        "due_date": [
            "next rd installment due date", "next rd installement due date",
            "due date", "next_due_date", "due_date"
        ],
    }

    @staticmethod
    def _find_column(df_columns: List[str], aliases: List[str]) -> Optional[str]:
        """Find a column by trying multiple name variations."""
        lower_cols = {c.lower().strip(): c for c in df_columns}
        for alias in aliases:
            if alias in lower_cols:
                return lower_cols[alias]
        return None

    @staticmethod
    def _parse_denomination(value) -> float:
        """Parse denomination from various formats.

        Examples:
            "2,000.00 Cr." → 2000.0
            "₹1,500"       → 1500.0
            1000            → 1000.0
            NaN             → 0.0
        """
        if pd.isna(value):
            return 0.0
        s = str(value)
        # Remove currency symbols, commas, "Cr.", spaces
        for char in ["₹", ",", "Cr.", "cr.", "CR.", " "]:
            s = s.replace(char, "")
        try:
            return float(s)
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _parse_date(value) -> Optional[date]:
        """Parse date from various formats."""
        if pd.isna(value):
            return None
        if isinstance(value, (datetime, pd.Timestamp)):
            return value.date()
        # Try common formats
        for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"]:
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except ValueError:
                continue
        return None

    @classmethod
    def load_from_excel(cls, filepath: str) -> Tuple[List[RDAccount], List[str]]:
        """Load RD accounts from an Excel file.

        Args:
            filepath: Path to the Excel file

        Returns:
            Tuple of (list of RDAccount objects, list of warning messages)

        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If required columns are missing
        """
        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {filepath}")

        # Read Excel
        df = pd.read_excel(filepath)
        warnings = []

        # Drop completely empty rows
        df = df.dropna(how="all").reset_index(drop=True)

        if df.empty:
            raise ValueError("Excel file contains no data rows.")

        # Map columns
        col_map = {}
        for field_name, aliases in cls.COLUMN_MAP.items():
            found = cls._find_column(df.columns.tolist(), aliases)
            if found:
                col_map[field_name] = found
            else:
                if field_name in ["account_no", "name", "denomination"]:
                    raise ValueError(
                        f"Required column '{field_name}' not found. "
                        f"Tried: {aliases}. Available: {df.columns.tolist()}"
                    )
                warnings.append(f"Optional column '{field_name}' not found.")

        # Parse accounts
        accounts = []
        seen_accounts = set()

        for idx, row in df.iterrows():
            # Skip rows with no account number
            raw_acc = row.get(col_map.get("account_no", ""), None)
            if pd.isna(raw_acc):
                continue

            account_no = str(int(raw_acc)) if isinstance(raw_acc, float) else str(raw_acc).strip()

            # Duplicate check
            if account_no in seen_accounts:
                warnings.append(f"Row {idx + 2}: Duplicate account '{account_no}' skipped.")
                continue
            seen_accounts.add(account_no)

            # Parse fields
            name = str(row.get(col_map.get("name", ""), "Unknown")).strip()
            if name == "nan" or not name:
                name = "Unknown"

            denomination = cls._parse_denomination(
                row.get(col_map.get("denomination", ""), 0)
            )
            if denomination <= 0:
                warnings.append(f"Row {idx + 2}: Invalid denomination for '{account_no}', skipping.")
                continue

            months_paid = 0
            if "months_paid" in col_map:
                mp = row.get(col_map["months_paid"], 0)
                months_paid = int(mp) if pd.notna(mp) else 0

            due_date = None
            if "due_date" in col_map:
                due_date = cls._parse_date(row.get(col_map["due_date"]))

            account = RDAccount(
                account_no=account_no,
                name=name,
                denomination=denomination,
                months_paid=months_paid,
                due_date=due_date,
            )
            accounts.append(account)

        if not accounts:
            raise ValueError("No valid accounts found in the Excel file.")

        return accounts, warnings

    @classmethod
    def load_from_pdf(cls, filepath: str) -> Tuple[List[RDAccount], List[str]]:
        """Load RD accounts from a PDF file containing a deposit accounts table.

        Extracts the table with columns: Account No, Account Name,
        Denomination, Month Paid Upto, Next RD Installment Due Date.

        Args:
            filepath: Path to the PDF file

        Returns:
            Tuple of (list of RDAccount objects, list of warning messages)
        """
        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"PDF file not found: {filepath}")

        warnings = []
        all_rows = []

        with pdfplumber.open(filepath) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables(
                    table_settings={"text_x_tolerance": 3, "text_y_tolerance": 3}
                )
                for table in tables:
                    if not table:
                        continue
                    for row in table:
                        if row:
                            all_rows.append(row)

        if not all_rows:
            raise ValueError("No tables found in the PDF file.")

        # Find header row by looking for known column keywords
        header_keywords = ["account no", "account name", "denomination", "month paid"]
        header_idx = None
        for i, row in enumerate(all_rows):
            row_text = " ".join(str(cell).lower().strip() for cell in row if cell)
            if any(kw in row_text for kw in header_keywords):
                header_idx = i
                break

        if header_idx is None:
            raise ValueError(
                "Could not find table header row in PDF. "
                "Expected columns: Account No, Account Name, Denomination, "
                "Month Paid Upto, Next RD Installment Due Date."
            )

        # Handle multi-line headers: join with space, normalize whitespace
        raw_headers = all_rows[header_idx]
        headers = []
        for h in raw_headers:
            if h:
                # Replace newlines within a cell with spaces
                headers.append(" ".join(str(h).split()))
            else:
                headers.append("")

        # Regex to parse a merged row where all columns are in one string
        # Matches: [select_num] account_no [: or space] name  denomination  months  date
        _merged_row_re = re.compile(
            r'(?:\d+\s+)?'                       # optional Select number
            r'(\d+)\s*:?\s*'                      # Account No
            r'(.+?)\s+'                           # Account Name (non-greedy)
            r'([\d,]+\.\d{2}\s*Cr\.?)\s+'         # Denomination e.g. "2,000.00 Cr."
            r'(\d+)\s+'                           # Month Paid Upto
            r'(\d{1,2}-[A-Za-z]{3}-\d{2,4})'     # Due Date e.g. "21-Jan-2026"
        )

        # Collect data rows, skipping repeated header rows on subsequent pages
        data_rows = []
        for row in all_rows[header_idx + 1:]:
            row_text = " ".join(str(cell).lower().strip() for cell in row if cell)
            # Skip rows that look like a repeated header
            if any(kw in row_text for kw in header_keywords):
                continue
            # Skip rows that are entirely empty or whitespace
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            # Detect merged rows: only 1 non-empty cell containing all data
            non_empty = [c for c in row if c is not None and str(c).strip()]
            if len(non_empty) == 1:
                merged_text = " ".join(str(non_empty[0]).split())
                m = _merged_row_re.search(merged_text)
                if m:
                    # Rebuild row to match header column count
                    # headers: [Select, Account No, Account Name, Denomination,
                    #           Month Paid Upto, Next RD Installment Due Date]
                    parsed = list(m.groups())  # 5 groups
                    # Pad to match header length (prepend empty Select column if needed)
                    while len(parsed) < len(headers):
                        parsed.insert(0, "")
                    row = parsed

            # Pad short rows to match header length
            if len(row) < len(headers):
                row = list(row) + [None] * (len(headers) - len(row))
            data_rows.append(row)

        if not data_rows:
            raise ValueError("No data rows found after the header in the PDF.")

        # Build DataFrame from extracted rows
        df = pd.DataFrame(data_rows, columns=headers)
        df = df.dropna(how="all").reset_index(drop=True)

        # Map columns using existing COLUMN_MAP logic
        col_map = {}
        for field_name, aliases in cls.COLUMN_MAP.items():
            found = cls._find_column(df.columns.tolist(), aliases)
            if found:
                col_map[field_name] = found
            else:
                if field_name in ["account_no", "name", "denomination"]:
                    raise ValueError(
                        f"Required column '{field_name}' not found in PDF table. "
                        f"Tried: {aliases}. Available: {df.columns.tolist()}"
                    )
                warnings.append(f"Optional column '{field_name}' not found in PDF.")

        # Parse accounts (same logic as load_from_excel)
        accounts = []
        seen_accounts = set()

        for idx, row in df.iterrows():
            raw_acc = row.get(col_map.get("account_no", ""), None)
            if pd.isna(raw_acc) or str(raw_acc).strip() == "":
                continue

            account_no = str(raw_acc).strip()
            # Remove .0 from float-like strings
            if account_no.endswith(".0"):
                account_no = account_no[:-2]

            if account_no in seen_accounts:
                warnings.append(f"PDF row {idx + 1}: Duplicate account '{account_no}' skipped.")
                continue
            seen_accounts.add(account_no)

            name = str(row.get(col_map.get("name", ""), "Unknown")).strip()
            if name == "nan" or name == "None" or not name:
                name = "Unknown"

            denomination = cls._parse_denomination(
                row.get(col_map.get("denomination", ""), 0)
            )
            if denomination <= 0:
                warnings.append(f"PDF row {idx + 1}: Invalid denomination for '{account_no}', skipping.")
                continue

            months_paid = 0
            if "months_paid" in col_map:
                mp = row.get(col_map["months_paid"], 0)
                if pd.notna(mp):
                    try:
                        months_paid = int(float(str(mp).strip()))
                    except (ValueError, TypeError):
                        months_paid = 0

            due_date = None
            if "due_date" in col_map:
                raw_date = row.get(col_map["due_date"])
                if raw_date and str(raw_date).strip():
                    due_date = cls._parse_date(raw_date)
                    # Also try "21-Jan-2026" format common in post office PDFs
                    if due_date is None:
                        for fmt in ["%d-%b-%Y", "%d %b %Y", "%d-%b-%y"]:
                            try:
                                due_date = datetime.strptime(str(raw_date).strip(), fmt).date()
                                break
                            except ValueError:
                                continue

            account = RDAccount(
                account_no=account_no,
                name=name,
                denomination=denomination,
                months_paid=months_paid,
                due_date=due_date,
            )
            accounts.append(account)

        if not accounts:
            raise ValueError("No valid accounts found in the PDF file.")

        return accounts, warnings


# ═══════════════════════════════════════════════════════════════
# SECTION 3: BUSINESS LOGIC - DEPOSIT LIST GENERATOR
# ═══════════════════════════════════════════════════════════════

class DepositListGenerator:
    """Generates optimized deposit lists from paid accounts.

    Algorithm:
        1. Filter only paid, active (non-matured) accounts
        2. Sort: HIGH priority first, then by denomination (descending)
        3. Greedy knapsack: fill each list up to max_amount
        4. Prefer round totals when possible

    This handles 1000+ accounts efficiently (O(n log n) sort + O(n) pass).
    """

    @staticmethod
    def generate(
        accounts: List[RDAccount],
        max_per_list: float = 20000,
    ) -> List[DepositList]:
        """Generate deposit lists from paid accounts.

        Args:
            accounts: All accounts (will filter for paid + active)
            max_per_list: Maximum amount per deposit list

        Returns:
            List of DepositList objects, each within the max amount
        """
        # Filter: only paid, non-expired accounts (matured accounts can be extended)
        eligible = [
            a for a in accounts
            if a.is_paid_this_month and a.status != AccountStatus.EXPIRED
        ]

        if not eligible:
            return []

        # Sort: HIGH priority first, then denomination descending
        eligible.sort(
            key=lambda a: (
                0 if a.priority == Priority.HIGH else 1,
                -a.denomination,
            )
        )

        lists = []
        used = set()

        while len(used) < len(eligible):
            current = DepositList(max_amount=max_per_list)
            added_any = False

            for acc in eligible:
                if acc.account_no in used:
                    continue
                if current.can_add(acc):
                    current.add(acc)
                    used.add(acc.account_no)
                    added_any = True

            if not added_any:
                break

            lists.append(current)

        return lists


# ═══════════════════════════════════════════════════════════════
# SECTION 4: ACCOUNT MANAGER (Central Business Service)
# ═══════════════════════════════════════════════════════════════

class AccountManager:
    """Central service for managing RD accounts.

    Provides:
        - Payment marking (paid/unpaid toggle)
        - Search and filtering
        - Statistics computation
        - Deposit list generation
    """

    def __init__(self, accounts: List[RDAccount] = None):
        self.accounts: List[RDAccount] = accounts or []

    def load_from_excel(self, filepath: str) -> List[str]:
        """Load accounts from Excel file. Returns warnings."""
        self.accounts, warnings = DataLoader.load_from_excel(filepath)
        return warnings

    def load_from_pdf(self, filepath: str) -> List[str]:
        """Load accounts from PDF file. Returns warnings."""
        self.accounts, warnings = DataLoader.load_from_pdf(filepath)
        return warnings

    def load_file(self, filepath: str) -> List[str]:
        """Load accounts from Excel or PDF based on file extension."""
        if filepath.lower().endswith(".pdf"):
            return self.load_from_pdf(filepath)
        return self.load_from_excel(filepath)

    @property
    def active_accounts(self) -> List[RDAccount]:
        """Non-expired accounts (includes matured accounts that are extended)."""
        return [a for a in self.accounts if a.status != AccountStatus.EXPIRED]

    @property
    def matured_accounts(self) -> List[RDAccount]:
        return [a for a in self.accounts if a.status == AccountStatus.MATURED]

    @property
    def expired_accounts(self) -> List[RDAccount]:
        """Accounts overdue by more than 6 months — cannot be continued."""
        return [a for a in self.accounts if a.status == AccountStatus.EXPIRED]

    @property
    def paid_accounts(self) -> List[RDAccount]:
        return [a for a in self.accounts if a.is_paid_this_month]

    @property
    def unpaid_accounts(self) -> List[RDAccount]:
        return [a for a in self.active_accounts if not a.is_paid_this_month]

    @property
    def defaulted_accounts(self) -> List[RDAccount]:
        return [a for a in self.accounts if a.has_default]

    @property
    def high_priority_accounts(self) -> List[RDAccount]:
        return [a for a in self.active_accounts if a.priority == Priority.HIGH]

    def mark_paid(self, account_no: str) -> bool:
        """Mark an account as paid for current month."""
        for acc in self.accounts:
            if acc.account_no == account_no:
                if acc.status == AccountStatus.EXPIRED:
                    return False
                acc.is_paid_this_month = True
                return True
        return False

    def mark_unpaid(self, account_no: str) -> bool:
        """Mark an account as unpaid (undo payment)."""
        for acc in self.accounts:
            if acc.account_no == account_no:
                acc.is_paid_this_month = False
                return True
        return False

    def toggle_paid(self, account_no: str) -> Optional[bool]:
        """Toggle paid status. Returns new status or None if not found."""
        for acc in self.accounts:
            if acc.account_no == account_no:
                if acc.status == AccountStatus.EXPIRED:
                    return None
                acc.is_paid_this_month = not acc.is_paid_this_month
                return acc.is_paid_this_month
        return None

    def mark_all_active_paid(self) -> int:
        """Mark all active accounts as paid. Returns count."""
        count = 0
        for acc in self.active_accounts:
            if not acc.is_paid_this_month:
                acc.is_paid_this_month = True
                count += 1
        return count

    def search(self, query: str) -> List[RDAccount]:
        """Search accounts by name or account number."""
        q = query.lower().strip()
        if not q:
            return self.accounts
        return [
            a for a in self.accounts
            if q in a.name.lower() or q in a.account_no.lower()
        ]

    def get_statistics(self) -> Dict:
        """Compute dashboard statistics."""
        paid_amt = sum(a.denomination for a in self.paid_accounts)
        return {
            "total_accounts": len(self.accounts),
            "active": len(self.active_accounts),
            "matured": len(self.matured_accounts),
            "expired": len(self.expired_accounts),
            "paid_this_month": len(self.paid_accounts),
            "unpaid": len(self.unpaid_accounts),
            "defaulters": len(self.defaulted_accounts),
            "high_priority": len(self.high_priority_accounts),
            "total_paid_amount": paid_amt,
        }

    def generate_deposit_lists(self, max_per_list: float = 20000) -> List[DepositList]:
        """Generate optimized deposit lists from paid accounts."""
        return DepositListGenerator.generate(self.accounts, max_per_list)

    @property
    def collected_accounts(self) -> List[RDAccount]:
        """Accounts where money is collected by agent but not yet deposited."""
        return self.paid_accounts

    def mark_collected(self, account_no: str) -> bool:
        """Mark account as collected (money received by agent)."""
        return self.mark_paid(account_no)

    def unmark_collected(self, account_no: str) -> bool:
        """Unmark collected status."""
        return self.mark_unpaid(account_no)


# ═══════════════════════════════════════════════════════════════
# SECTION 5: EXCEL EXPORTER (Formatted Output)
# ═══════════════════════════════════════════════════════════════

class ExcelExporter:
    """Generates professionally formatted Excel reports.

    Output includes:
        Sheet 1: Deposit Lists (grouped, with totals)
        Sheet 2: Payment Summary (statistics)
        Sheet 3: Full Account Data (for reference)
    """

    # Style constants
    HEADER_FILL = PatternFill("solid", fgColor="1a2540")
    HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ACCENT_FILL = PatternFill("solid", fgColor="FFF3E0")
    GREEN_FILL = PatternFill("solid", fgColor="E8F5E9")
    RED_FILL = PatternFill("solid", fgColor="FFEBEE")
    PURPLE_FILL = PatternFill("solid", fgColor="F3E5F5")
    GRAY_FILL = PatternFill("solid", fgColor="E0E0E0")  # Expired accounts
    TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1a2540")
    BOLD_FONT = Font(name="Arial", bold=True, size=11)
    NORMAL_FONT = Font(name="Arial", size=10)
    MONEY_FORMAT = '₹#,##0'
    THIN_BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    @classmethod
    def export_deposit_lists(
        cls,
        deposit_lists: List[DepositList],
        manager: AccountManager,
        output_path: str,
    ) -> str:
        """Export deposit lists and summary to a formatted Excel file.

        Args:
            deposit_lists: Generated deposit lists
            manager: AccountManager with all account data
            output_path: Where to save the Excel file

        Returns:
            The output file path
        """
        wb = Workbook()

        # ─── Sheet 1: Deposit Lists ───
        ws1 = wb.active
        ws1.title = "Deposit Lists"
        cls._write_deposit_sheet(ws1, deposit_lists)

        # ─── Sheet 2: Payment Summary ───
        ws2 = wb.create_sheet("Payment Summary")
        cls._write_summary_sheet(ws2, manager)

        # ─── Sheet 3: All Accounts ───
        ws3 = wb.create_sheet("All Accounts")
        cls._write_accounts_sheet(ws3, manager.accounts)

        # Save
        wb.save(output_path)
        return output_path

    @classmethod
    def _write_deposit_sheet(cls, ws, deposit_lists: List[DepositList]):
        """Write deposit lists to worksheet."""
        ws.sheet_properties.tabColor = "F59E0B"

        # Title
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "POST OFFICE RD - DEPOSIT LIST REPORT"
        title_cell.font = cls.TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center")

        ws["A2"] = f"Generated: {date.today().strftime('%d-%b-%Y')}"
        ws["A2"].font = Font(name="Arial", size=9, color="888888")

        row = 4
        grand_total = 0

        for idx, dlist in enumerate(deposit_lists, 1):
            # List header
            ws.merge_cells(f"A{row}:F{row}")
            header = ws.cell(row=row, column=1)
            header.value = f"DEPOSIT LIST {idx}  •  Total: ₹{dlist.total:,.0f}  •  {dlist.count} accounts"
            header.font = Font(name="Arial", bold=True, size=12, color="1a2540")
            header.fill = cls.ACCENT_FILL
            row += 1

            # Column headers
            headers = ["Sl No", "Account No", "Customer Name", "Amount (₹)", "Priority", "Due Date"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = cls.HEADER_FONT
                cell.fill = cls.HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
                cell.border = cls.THIN_BORDER
            row += 1

            # Account rows
            for i, acc in enumerate(dlist.accounts, 1):
                ws.cell(row=row, column=1, value=i).font = cls.NORMAL_FONT
                ws.cell(row=row, column=2, value=acc.account_no).font = cls.NORMAL_FONT
                ws.cell(row=row, column=3, value=acc.name).font = cls.NORMAL_FONT

                amt_cell = ws.cell(row=row, column=4, value=acc.denomination)
                amt_cell.font = cls.NORMAL_FONT
                amt_cell.number_format = cls.MONEY_FORMAT

                pri_cell = ws.cell(row=row, column=5, value=acc.priority.value)
                pri_cell.font = cls.NORMAL_FONT
                if acc.priority == Priority.HIGH:
                    pri_cell.fill = PatternFill("solid", fgColor="FFF8E1")

                due_cell = ws.cell(
                    row=row, column=6,
                    value=acc.due_date.strftime("%d-%b-%Y") if acc.due_date else "—"
                )
                due_cell.font = cls.NORMAL_FONT

                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = cls.THIN_BORDER
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

                row += 1

            # List total
            ws.cell(row=row, column=3, value="LIST TOTAL").font = cls.BOLD_FONT
            total_cell = ws.cell(row=row, column=4, value=dlist.total)
            total_cell.font = Font(name="Arial", bold=True, size=12, color="10B981")
            total_cell.number_format = cls.MONEY_FORMAT
            total_cell.border = Border(top=Side(style="double"))

            grand_total += dlist.total
            row += 2

        # Grand total
        ws.merge_cells(f"A{row}:C{row}")
        ws.cell(row=row, column=1, value="GRAND TOTAL").font = Font(
            name="Arial", bold=True, size=14, color="1a2540"
        )
        gt_cell = ws.cell(row=row, column=4, value=grand_total)
        gt_cell.font = Font(name="Arial", bold=True, size=14, color="10B981")
        gt_cell.number_format = cls.MONEY_FORMAT

        # Column widths
        widths = [8, 15, 25, 15, 12, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    @classmethod
    def _write_summary_sheet(cls, ws, manager: AccountManager):
        """Write payment summary statistics."""
        ws.sheet_properties.tabColor = "3B82F6"
        stats = manager.get_statistics()

        ws.merge_cells("A1:C1")
        ws["A1"].value = "PAYMENT SUMMARY"
        ws["A1"].font = cls.TITLE_FONT

        ws["A2"] = f"Date: {date.today().strftime('%d-%b-%Y')}"
        ws["A2"].font = Font(name="Arial", size=9, color="888888")

        row = 4
        items = [
            ("Total Accounts", stats["total_accounts"], None),
            ("Active Accounts", stats["active"], cls.GREEN_FILL),
            ("Matured (60+ months, extended)", stats["matured"], cls.PURPLE_FILL),
            ("Expired (>6 months overdue)", stats["expired"], cls.GRAY_FILL),
            ("Paid This Month", stats["paid_this_month"], cls.GREEN_FILL),
            ("Unpaid This Month", stats["unpaid"], cls.ACCENT_FILL),
            ("Defaulters (Overdue)", stats["defaulters"], cls.RED_FILL),
            ("High Priority (≤15th)", stats["high_priority"], cls.ACCENT_FILL),
        ]

        # Headers
        for col, h in enumerate(["Metric", "Count"], 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.border = cls.THIN_BORDER
        row += 1

        for label, value, fill in items:
            ws.cell(row=row, column=1, value=label).font = cls.NORMAL_FONT
            ws.cell(row=row, column=2, value=value).font = cls.BOLD_FONT
            if fill:
                ws.cell(row=row, column=1).fill = fill
                ws.cell(row=row, column=2).fill = fill
            for col in range(1, 3):
                ws.cell(row=row, column=col).border = cls.THIN_BORDER
            row += 1

        # Total paid amount
        row += 1
        ws.cell(row=row, column=1, value="Total Paid Amount").font = cls.BOLD_FONT
        amt = ws.cell(row=row, column=2, value=stats["total_paid_amount"])
        amt.font = Font(name="Arial", bold=True, size=14, color="10B981")
        amt.number_format = cls.MONEY_FORMAT

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18

    @classmethod
    def _write_accounts_sheet(cls, ws, accounts: List[RDAccount]):
        """Write full account data for reference."""
        ws.sheet_properties.tabColor = "A855F7"

        headers = [
            "Account No", "Name", "Denomination", "Months Paid",
            "Remaining", "Due Date", "Priority", "Status", "Paid?"
        ]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.border = cls.THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        for row_idx, acc in enumerate(accounts, 2):
            ws.cell(row=row_idx, column=1, value=acc.account_no)
            ws.cell(row=row_idx, column=2, value=acc.name)

            denom_cell = ws.cell(row=row_idx, column=3, value=acc.denomination)
            denom_cell.number_format = cls.MONEY_FORMAT

            ws.cell(row=row_idx, column=4, value=acc.months_paid)
            ws.cell(row=row_idx, column=5, value=acc.remaining_months)

            ws.cell(
                row=row_idx, column=6,
                value=acc.due_date.strftime("%d-%b-%Y") if acc.due_date else "—"
            )
            ws.cell(row=row_idx, column=7, value=acc.priority.value)
            ws.cell(row=row_idx, column=8, value=acc.status.value)
            ws.cell(row=row_idx, column=9, value="YES" if acc.is_paid_this_month else "NO")

            # Color coding
            status_cell = ws.cell(row=row_idx, column=8)
            if acc.status == AccountStatus.EXPIRED:
                status_cell.fill = cls.GRAY_FILL
            elif acc.status == AccountStatus.MATURED:
                status_cell.fill = cls.PURPLE_FILL
            elif acc.status == AccountStatus.DEFAULTED:
                status_cell.fill = cls.RED_FILL
            else:
                status_cell.fill = cls.GREEN_FILL

            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).font = cls.NORMAL_FONT
                ws.cell(row=row_idx, column=col).border = cls.THIN_BORDER
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")

        widths = [14, 22, 14, 12, 12, 14, 12, 14, 8]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════
# SECTION 6: CLI INTERFACE
# ═══════════════════════════════════════════════════════════════

class CLIInterface:
    """Command-line interface for the RD Schedule System.

    Provides a simple, menu-driven interface suitable for
    non-technical RD agents working on any computer.
    """

    def __init__(self):
        self.manager = AccountManager()
        self.deposit_lists = []

    def clear_screen(self):
        os.system("cls" if os.name == "nt" else "clear")

    def print_header(self, title: str):
        width = 60
        print("\n" + "═" * width)
        print(f"  {title}")
        print("═" * width)

    def print_divider(self):
        print("─" * 60)

    def run(self):
        """Main application loop."""
        self.clear_screen()
        print("""
╔══════════════════════════════════════════════════════════╗
║       POST OFFICE RD SCHEDULE AUTOMATION SYSTEM         ║
║       ─────────────────────────────────────────         ║
║       Version 1.0  •  Offline  •  1000+ Accounts        ║
╚══════════════════════════════════════════════════════════╝
        """)

        # Step 1: Load Excel or PDF
        filepath = input("📂 Enter Excel/PDF file path (or press Enter for demo): ").strip()
        if filepath:
            try:
                warnings = self.manager.load_file(filepath)
                print(f"\n✅ Loaded {len(self.manager.accounts)} accounts.")
                if warnings:
                    print(f"⚠️  {len(warnings)} warning(s):")
                    for w in warnings[:5]:
                        print(f"   • {w}")
            except Exception as e:
                print(f"\n❌ Error: {e}")
                print("Loading demo data instead...")
                self._load_demo_data()
        else:
            self._load_demo_data()

        # Main menu loop
        while True:
            self._show_menu()
            choice = input("\n👉 Enter choice (1-7): ").strip()

            if choice == "1":
                self._show_dashboard()
            elif choice == "2":
                self._payment_tracking()
            elif choice == "3":
                self._mark_payments()
            elif choice == "4":
                self._generate_deposit_list()
            elif choice == "5":
                self._export_excel()
            elif choice == "6":
                self._search_accounts()
            elif choice == "7":
                print("\n👋 Thank you for using RD Schedule System. Goodbye!")
                break
            else:
                print("❌ Invalid choice. Try again.")

    def _show_menu(self):
        self.print_header("MAIN MENU")
        print("""
  1. 📊 Dashboard (Overview)
  2. 💰 View Payment Status
  3. ✅ Mark Payments (Paid/Unpaid)
  4. 📋 Generate Deposit List
  5. 📥 Export to Excel
  6. 🔍 Search Accounts
  7. 🚪 Exit
        """)

    def _show_dashboard(self):
        stats = self.manager.get_statistics()
        self.print_header("DASHBOARD")
        print(f"""
  📊 Total Accounts    : {stats['total_accounts']}
  ✅ Active            : {stats['active']}
  🏆 Matured (60m+)    : {stats['matured']}
  🚫 Expired (>6m due) : {stats['expired']}
  💰 Paid This Month   : {stats['paid_this_month']}
  ⏳ Unpaid            : {stats['unpaid']}
  ⚠️  Defaulters        : {stats['defaulters']}
  ⭐ High Priority     : {stats['high_priority']}

  💵 Total Paid Amount : ₹{stats['total_paid_amount']:,.0f}
        """)
        input("Press Enter to continue...")

    def _payment_tracking(self):
        self.print_header("PAYMENT STATUS")
        print(f"\n{'No':<4} {'Account':<10} {'Name':<20} {'Amount':>10} {'Status':<10} {'Priority':<8}")
        self.print_divider()

        for i, acc in enumerate(self.manager.accounts[:30], 1):
            if acc.is_expired:
                status = "🚫 EXPIRED"
            elif acc.is_matured:
                status = "MATURED"
            elif acc.is_overdue and not acc.is_paid_this_month:
                status = "⚠️ DEFAULT"
            elif acc.is_paid_this_month:
                status = "✅ PAID"
            else:
                status = "❌ UNPAID"
            print(
                f"{i:<4} {acc.account_no:<10} {acc.name:<20} "
                f"₹{acc.denomination:>8,.0f} {status:<10} {acc.priority.value:<8}"
            )

        if len(self.manager.accounts) > 30:
            print(f"\n  ... and {len(self.manager.accounts) - 30} more accounts")
        input("\nPress Enter to continue...")

    def _mark_payments(self):
        self.print_header("MARK PAYMENTS")
        print("\nOptions:")
        print("  1. Mark individual account as Paid")
        print("  2. Mark ALL active accounts as Paid")
        print("  3. Toggle account (Paid ↔ Unpaid)")
        print("  4. Back to menu")

        choice = input("\n👉 Choice: ").strip()

        if choice == "1":
            acc_no = input("Enter Account No: ").strip()
            if self.manager.mark_paid(acc_no):
                print(f"✅ Account {acc_no} marked as PAID.")
            else:
                print(f"❌ Account {acc_no} not found or expired.")

        elif choice == "2":
            count = self.manager.mark_all_active_paid()
            print(f"✅ Marked {count} accounts as PAID.")

        elif choice == "3":
            acc_no = input("Enter Account No: ").strip()
            result = self.manager.toggle_paid(acc_no)
            if result is not None:
                status = "PAID" if result else "UNPAID"
                print(f"✅ Account {acc_no} → {status}")
            else:
                print(f"❌ Account {acc_no} not found or expired.")

        input("\nPress Enter to continue...")

    def _generate_deposit_list(self):
        self.print_header("GENERATE DEPOSIT LIST")

        max_amt = input("Max amount per list (default ₹20,000): ").strip()
        max_amt = float(max_amt) if max_amt else 20000

        self.deposit_lists = self.manager.generate_deposit_lists(max_amt)

        if not self.deposit_lists:
            print("\n❌ No paid accounts found. Mark payments first (Option 3).")
            input("Press Enter to continue...")
            return

        print(f"\n✅ Generated {len(self.deposit_lists)} deposit list(s):\n")

        grand = 0
        for idx, dlist in enumerate(self.deposit_lists, 1):
            print(f"  📋 List {idx}: {dlist.count} accounts • ₹{dlist.total:,.0f}")
            for acc in dlist.accounts:
                print(f"     {acc.account_no:<10} {acc.name:<20} ₹{acc.denomination:>8,.0f}  [{acc.priority.value}]")
            self.print_divider()
            grand += dlist.total

        print(f"\n  💰 GRAND TOTAL: ₹{grand:,.0f}")
        input("\nPress Enter to continue...")

    def _export_excel(self):
        self.print_header("EXPORT TO EXCEL")

        if not self.deposit_lists:
            print("\n⚠️  No deposit lists generated yet. Generating now...")
            self.deposit_lists = self.manager.generate_deposit_lists()
            if not self.deposit_lists:
                print("❌ No paid accounts. Mark payments first.")
                input("Press Enter to continue...")
                return

        output = input("Output filename (default: RD_Deposit_Report.xlsx): ").strip()
        if not output:
            output = f"RD_Deposit_Report_{date.today().strftime('%Y%m%d')}.xlsx"
        if not output.endswith(".xlsx"):
            output += ".xlsx"

        try:
            path = ExcelExporter.export_deposit_lists(
                self.deposit_lists, self.manager, output
            )
            print(f"\n✅ Report exported to: {os.path.abspath(path)}")
            print("   Sheets: Deposit Lists | Payment Summary | All Accounts")
        except Exception as e:
            print(f"\n❌ Export failed: {e}")

        input("\nPress Enter to continue...")

    def _search_accounts(self):
        self.print_header("SEARCH ACCOUNTS")
        query = input("🔍 Enter name or account no: ").strip()
        results = self.manager.search(query)

        if not results:
            print("No results found.")
        else:
            print(f"\nFound {len(results)} result(s):\n")
            for acc in results[:20]:
                status = acc.status.value
                paid = "PAID" if acc.is_paid_this_month else "UNPAID"
                print(
                    f"  {acc.account_no:<10} {acc.name:<20} "
                    f"₹{acc.denomination:>8,.0f}  {acc.months_paid}m  "
                    f"{status:<10} {paid}"
                )

        input("\nPress Enter to continue...")

    def _load_demo_data(self):
        """Load demonstration data for testing."""
        import random
        random.seed(42)

        names = [
            "Raman K", "Seetha R", "Lakshmanan P", "Priya M", "Suresh B",
            "Anitha S", "Gopal V", "Meena D", "Vijay K", "Deepa L",
            "Karthik N", "Saroja T", "Murugan A", "Lakshmi G", "Ravi S",
            "Padma H", "Senthil R", "Valli M", "Babu C", "Chitra J",
        ]
        denoms = [500, 1000, 1500, 2000, 3000]

        accounts = []
        for i in range(20):
            day = random.randint(1, 28)
            month = random.choice([1, 2, 3, 4])
            months_paid = random.choice([10, 20, 30, 42, 45, 48, 50, 55, 58, 59, 60, 62])

            acc = RDAccount(
                account_no=str(1001 + i),
                name=names[i % len(names)],
                denomination=random.choice(denoms),
                months_paid=months_paid,
                due_date=date(2026, month, day),
            )
            accounts.append(acc)

        self.manager = AccountManager(accounts)
        print(f"✅ Loaded {len(accounts)} demo accounts.")


# ═══════════════════════════════════════════════════════════════
# SECTION 7: GUI INTERFACE
# ═══════════════════════════════════════════════════════════════

class GUIInterface:
    """Desktop GUI interface for RD Schedule System."""

    def __init__(self):
        self.manager = AccountManager()
        self.deposit_lists: List[DepositList] = []
        self.account_repeat_map: Dict[str, int] = {}
        self.root = tk.Tk()
        self.root.title("Post Office RD Schedule System")
        self.root.geometry("1200x760")
        self.root.minsize(980, 620)
        self._setup_styles()
        self._build_ui()
        self._refresh_all()

    def _setup_styles(self):
        """Apply a cleaner visual style to the application."""
        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        self.root.configure(bg="#f4f7fb")
        style.configure("TFrame", background="#f4f7fb")
        style.configure("TLabel", background="#f4f7fb", font=("Segoe UI", 10))
        style.configure("Header.TLabel", background="#1f3b73", foreground="#ffffff", font=("Segoe UI Semibold", 13))
        style.configure("Stats.TLabel", background="#e8eef9", foreground="#1a2540", font=("Segoe UI", 10))
        style.configure("TButton", padding=(10, 6), font=("Segoe UI", 9))
        style.configure("Accent.TButton", padding=(11, 6), font=("Segoe UI Semibold", 9))
        style.configure("TEntry", padding=4)
        style.configure("Treeview.Heading", font=("Segoe UI Semibold", 9))
        style.configure("Treeview", rowheight=24, font=("Segoe UI", 9))
        style.configure("TLabelframe", background="#f4f7fb")
        style.configure("TLabelframe.Label", background="#f4f7fb", font=("Segoe UI Semibold", 10))

    def _build_ui(self):
        title_bar = ttk.Label(
            self.root,
            text="POST OFFICE RD SCHEDULE SYSTEM",
            style="Header.TLabel",
            anchor=tk.CENTER,
            padding=(12, 10),
        )
        title_bar.pack(fill=tk.X)

        top = ttk.Frame(self.root, padding=10)
        top.pack(fill=tk.X)

        ttk.Button(top, text="Load Excel/PDF", style="Accent.TButton", command=self._load_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(top, text="Collected Window", command=self._open_collected_window).pack(side=tk.LEFT, padx=5)
        ttk.Button(top, text="View Collected Accounts", command=self._open_collected_accounts_window).pack(side=tk.LEFT, padx=5)
        ttk.Button(top, text="Export Excel", style="Accent.TButton", command=self._export_excel).pack(side=tk.LEFT, padx=5)

        search_row = ttk.Frame(self.root, padding=(10, 0, 10, 8))
        search_row.pack(fill=tk.X)
        ttk.Label(search_row, text="Search Account:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_row, textvariable=self.search_var, width=45)
        search_entry.pack(side=tk.LEFT, padx=6)
        search_entry.bind("<KeyRelease>", lambda _e: self._refresh_table())
        ttk.Button(search_row, text="Clear", command=self._clear_search).pack(side=tk.LEFT)

        self.stats_var = tk.StringVar(value="No data loaded.")
        ttk.Label(
            self.root,
            textvariable=self.stats_var,
            style="Stats.TLabel",
            padding=(10, 8, 10, 8),
            anchor=tk.W,
        ).pack(fill=tk.X, padx=10, pady=(0, 8))

        table_frame = ttk.Frame(self.root, padding=(10, 0, 10, 8))
        table_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("account_no", "name", "amount", "months_paid", "status", "priority", "paid", "due_date")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=18)
        self.tree.heading("account_no", text="Account No")
        self.tree.heading("name", text="Name")
        self.tree.heading("amount", text="Amount")
        self.tree.heading("months_paid", text="Months Paid")
        self.tree.heading("status", text="Status")
        self.tree.heading("priority", text="Priority")
        self.tree.heading("paid", text="Paid")
        self.tree.heading("due_date", text="Due Date")

        self.tree.column("account_no", width=110, anchor=tk.CENTER)
        self.tree.column("name", width=240, anchor=tk.W)
        self.tree.column("amount", width=100, anchor=tk.E)
        self.tree.column("months_paid", width=100, anchor=tk.CENTER)
        self.tree.column("status", width=100, anchor=tk.CENTER)
        self.tree.column("priority", width=90, anchor=tk.CENTER)
        self.tree.column("paid", width=75, anchor=tk.CENTER)
        self.tree.column("due_date", width=110, anchor=tk.CENTER)
        self.tree.tag_configure("paid", background="#eaf8ec")
        self.tree.tag_configure("defaulted", background="#fdecec")
        self.tree.tag_configure("matured", background="#f3ecfd")
        self.tree.tag_configure("expired", background="#e0e0e0")

        yscroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)

        preview_wrap = ttk.LabelFrame(self.root, text="Deposit List Preview", padding=8)
        preview_wrap.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        self.preview_text = ScrolledText(preview_wrap, height=10)
        self.preview_text.configure(
            font=("Consolas", 10),
            bg="#ffffff",
            fg="#1a2540",
            padx=8,
            pady=8,
            relief=tk.FLAT,
        )
        self.preview_text.pack(fill=tk.BOTH, expand=True)

    def _clear_search(self):
        self.search_var.set("")
        self._refresh_table()

    def _refresh_stats(self):
        stats = self.manager.get_statistics()
        self.stats_var.set(
            f"Total: {stats['total_accounts']} | Active: {stats['active']} | "
            f"Matured: {stats['matured']} | Expired: {stats['expired']} | "
            f"Collected(Not Deposited): {stats['paid_this_month']} | "
            f"Unpaid: {stats['unpaid']} | Defaulters: {stats['defaulters']} | "
            f"High Priority: {stats['high_priority']} | Collected Amount: Rs {stats['total_paid_amount']:,.0f}"
        )

    def _refresh_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        query = self.search_var.get().strip()
        accounts = self.manager.search(query) if query else self.manager.accounts

        for acc in accounts:
            due = acc.due_date.strftime("%d-%b-%Y") if acc.due_date else "-"
            paid = "YES" if acc.is_paid_this_month else "NO"
            tag = ""
            if acc.status == AccountStatus.EXPIRED:
                tag = "expired"
            elif acc.status == AccountStatus.DEFAULTED:
                tag = "defaulted"
            elif acc.status == AccountStatus.MATURED:
                tag = "matured"
            elif acc.is_paid_this_month:
                tag = "paid"
            self.tree.insert(
                "",
                tk.END,
                values=(
                    acc.account_no,
                    acc.name,
                    f"{acc.denomination:,.0f}",
                    acc.months_paid,
                    acc.status.value,
                    acc.priority.value,
                    paid,
                    due,
                ),
                tags=(tag,) if tag else (),
            )

    def _refresh_preview(self):
        self.preview_text.delete("1.0", tk.END)
        if not self.deposit_lists:
            self.preview_text.insert(tk.END, "No deposit lists generated.")
            return

        grand_total = 0
        for idx, dlist in enumerate(self.deposit_lists, 1):
            self.preview_text.insert(
                tk.END,
                f"List {idx}: {dlist.count} accounts | Total Rs {dlist.total:,.0f}\n",
            )
            for acc in dlist.accounts:
                self.preview_text.insert(
                    tk.END,
                    f"  {acc.account_no:<10} {acc.name:<22} Rs {acc.denomination:>8,.0f} [{acc.priority.value}]\n",
                )
            self.preview_text.insert(tk.END, "-" * 70 + "\n")
            grand_total += dlist.total
        self.preview_text.insert(tk.END, f"GRAND TOTAL: Rs {grand_total:,.0f}\n")

    def _refresh_all(self):
        self._refresh_stats()
        self._refresh_table()
        self._refresh_preview()

    def _load_excel(self):
        filepath = filedialog.askopenfilename(
            title="Select account file (Excel or PDF)",
            filetypes=[
                ("Supported files", "*.xlsx *.xls *.pdf"),
                ("Excel files", "*.xlsx *.xls"),
                ("PDF files", "*.pdf"),
                ("All files", "*.*"),
            ],
        )
        if not filepath:
            return

        try:
            warnings = self.manager.load_file(filepath)
            self.deposit_lists = []
            self._refresh_all()
            message = f"Loaded {len(self.manager.accounts)} accounts."
            if warnings:
                message += f"\nWarnings: {len(warnings)}"
            messagebox.showinfo("Load Complete", message)
        except Exception as exc:
            messagebox.showerror("Load Failed", str(exc))

    def _load_demo_data(self):
        cli = CLIInterface()
        cli._load_demo_data()
        self.manager = cli.manager
        self.deposit_lists = []
        self._refresh_all()
        messagebox.showinfo("Demo Data", f"Loaded {len(self.manager.accounts)} demo accounts.")

    def _selected_account_no(self) -> Optional[str]:
        selected = self.tree.selection()
        if not selected:
            return None
        values = self.tree.item(selected[0], "values")
        if not values:
            return None
        return str(values[0]).strip()

    def _toggle_selected_paid(self):
        account_no = self._selected_account_no()
        if not account_no:
            messagebox.showwarning("No Selection", "Select one account from the table.")
            return
        result = self.manager.toggle_paid(account_no)
        if result is None:
            messagebox.showwarning("Cannot Update", "Account not found or already matured.")
            return
        self.deposit_lists = []
        self._refresh_all()

    def _mark_selected_collected(self):
        account_no = self._selected_account_no()
        if not account_no:
            messagebox.showwarning("No Selection", "Select one account from the table.")
            return
        if not self.manager.mark_collected(account_no):
            messagebox.showwarning("Cannot Update", "Account not found or already matured.")
            return
        self.deposit_lists = []
        self._refresh_all()

    def _unmark_selected_collected(self):
        account_no = self._selected_account_no()
        if not account_no:
            messagebox.showwarning("No Selection", "Select one account from the table.")
            return
        if not self.manager.unmark_collected(account_no):
            messagebox.showwarning("Cannot Update", "Account not found.")
            return
        self.deposit_lists = []
        self._refresh_all()

    def _mark_all_paid(self):
        count = self.manager.mark_all_active_paid()
        self.deposit_lists = []
        self._refresh_all()
        messagebox.showinfo("Done", f"Marked {count} accounts as paid.")

    def _open_collected_window(self):
        win = tk.Toplevel(self.root)
        win.title("Collected by Agent (Not Yet Deposited)")
        win.geometry("900x560")
        win.transient(self.root)
        win.grab_set()

        top = ttk.Frame(win, padding=8)
        top.pack(fill=tk.X)

        ttk.Label(top, text="Search:").pack(side=tk.LEFT)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(top, textvariable=search_var, width=40)
        search_entry.pack(side=tk.LEFT, padx=6)
        ttk.Label(top, text="Status Filter:").pack(side=tk.LEFT, padx=(10, 4))
        status_filter_var = tk.StringVar(value="All")
        status_filter = ttk.Combobox(
            top,
            textvariable=status_filter_var,
            values=["All", "Defaulted Only", "Non-defaulted"],
            width=16,
            state="readonly",
        )
        status_filter.pack(side=tk.LEFT)
        status_filter.current(0)

        ttk.Label(top, text="Due Date Filter:").pack(side=tk.LEFT, padx=(10, 4))
        due_filter_var = tk.StringVar(value="All")
        due_filter = ttk.Combobox(
            top,
            textvariable=due_filter_var,
            values=[
                "All",
                "Due <= 15 (This Month) Only",
                "Hide Due <= 15 (This Month)",
            ],
            width=28,
            state="readonly",
        )
        due_filter.pack(side=tk.LEFT)
        due_filter.current(0)

        table_wrap = ttk.Frame(win, padding=(8, 0, 8, 8))
        table_wrap.pack(fill=tk.BOTH, expand=True)

        checked_accounts = set()
        cols = ("select", "account_no", "name", "amount","months_missed","due_date", "status", "collected")
        tree = ttk.Treeview(table_wrap, columns=cols, show="headings")
        tree.heading("select", text="Select")
        tree.heading("account_no", text="Account No")
        tree.heading("name", text="Name")
        tree.heading("amount", text="Amount")
        tree.heading("months_missed", text="Months Missed")        
        tree.heading("due_date", text="Due Date")
        tree.heading("status", text="Status")
        tree.heading("collected", text="Collected?")

        tree.column("select", width=70, anchor=tk.CENTER)
        tree.column("account_no", width=120, anchor=tk.CENTER)
        tree.column("name", width=280, anchor=tk.W)
        tree.column("amount", width=120, anchor=tk.E)
        tree.column("months_missed", width=130, anchor=tk.CENTER)
        tree.column("due_date", width=120, anchor=tk.CENTER)
        tree.column("status", width=120, anchor=tk.CENTER)
        tree.column("collected", width=120, anchor=tk.CENTER)

        # --- Column sort state & logic ---
        sort_state: Dict[str, bool] = {}  # col -> True=ascending, False=descending

        # Original heading labels (without sort arrows)
        heading_labels = {
            "account_no": "Account No", "name": "Name", "amount": "Amount",
            "months_missed": "Months Missed", "due_date": "Due Date",
            "status": "Status", "collected": "Collected?",
        }

        def _sort_key(col: str, value: str):
            """Return a sort key appropriate for the column type."""
            val = str(value).strip()
            if col in ("amount",):
                # Numeric: strip commas and non-digit chars
                cleaned = val.replace(",", "").replace("Cr.", "").replace("Cr", "").strip()
                try:
                    return float(cleaned)
                except ValueError:
                    return 0.0
            if col in ("months_missed",):
                try:
                    return int(val)
                except ValueError:
                    return 0
            if col == "due_date":
                try:
                    return datetime.strptime(val, "%d-%b-%Y")
                except ValueError:
                    return datetime.min
            return val.lower()

        def _sort_by_column(col: str):
            """Sort treeview rows by the clicked column, toggling asc/desc."""
            ascending = not sort_state.get(col, False)  # toggle
            sort_state[col] = ascending

            # Gather all rows with their values and tags
            rows = []
            for iid in tree.get_children():
                vals = tree.item(iid, "values")
                tags = tree.item(iid, "tags")
                rows.append((iid, vals, tags))

            # Determine column index
            col_idx = cols.index(col)

            # Sort rows
            rows.sort(
                key=lambda r: _sort_key(col, r[1][col_idx] if col_idx < len(r[1]) else ""),
                reverse=not ascending,
            )

            # Rearrange treeview rows in sorted order
            for idx, (iid, _vals, _tags) in enumerate(rows):
                tree.move(iid, "", idx)

            # Update heading text with sort arrow indicator
            arrow = " ▲" if ascending else " ▼"
            for c, label in heading_labels.items():
                if c == col:
                    tree.heading(c, text=label + arrow)
                else:
                    tree.heading(c, text=label)

        # Bind column heading click for all sortable columns (skip "select")
        for col_name in heading_labels:
            tree.heading(
                col_name,
                text=heading_labels[col_name],
                command=lambda c=col_name: _sort_by_column(c),
            )

        yscroll = ttk.Scrollbar(table_wrap, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)

        tree.tag_configure("due_le_15_this_month", background="#FEF3C7")

        def refresh_popup_table():
            for row in tree.get_children():
                tree.delete(row)
            q = search_var.get().strip().lower()
            selected_filter = status_filter_var.get().strip()
            today = date.today()
            due_filter_choice = due_filter_var.get().strip()

            def due_le_15_this_month_for(acc: RDAccount) -> bool:
                return (
                    acc.due_date
                    and acc.due_date.year == today.year
                    and acc.due_date.month == today.month
                    and acc.due_date.day <= 15
                )

            def months_missed_for(acc: RDAccount) -> int:
                """Estimate missed installments count based on how many months overdue."""
                if not acc.due_date:
                    return 0
                if acc.due_date >= today:
                    return 0
                month_diff = (today.year * 12 + today.month) - (acc.due_date.year * 12 + acc.due_date.month)
                # If overdue within same month (due date earlier than today), count 1.
                return max(1, month_diff)

            for acc in self.manager.active_accounts:
                if q and q not in acc.account_no.lower() and q not in acc.name.lower():
                    continue
                if selected_filter == "Defaulted Only" and not acc.has_default:
                    continue
                if selected_filter == "Non-defaulted" and acc.has_default:
                    continue

                due_flag = due_le_15_this_month_for(acc)
                if due_filter_choice == "Due <= 15 (This Month) Only" and not due_flag:
                    continue
                if due_filter_choice == "Hide Due <= 15 (This Month)" and due_flag:
                    continue

                tree.insert(
                    "",
                    tk.END,
                    values=(
                        "[x]" if acc.account_no in checked_accounts else "[ ]",
                        acc.account_no,
                        acc.name,
                        f"{acc.denomination:,.0f}",
                        months_missed_for(acc),
                        acc.due_date.strftime("%d-%b-%Y") if acc.due_date else "-", 
                        acc.status.value,
                        "YES" if acc.is_paid_this_month else "NO",
                        
                        
                    ),
                    tags=("due_le_15_this_month",) if due_flag else (),
                )

        def selected_account():
            selected = tree.selection()
            if not selected:
                return None
            vals = tree.item(selected[0], "values")
            return str(vals[1]).strip() if vals else None

        def _selected_checked_accounts() -> List[str]:
            selected = []
            for row in tree.get_children():
                vals = tree.item(row, "values")
                if vals and str(vals[0]).strip() == "[x]":
                    selected.append(str(vals[1]).strip())
            return selected

        def toggle_row_checkbox(event):
            row_id = tree.identify_row(event.y)
            col_id = tree.identify_column(event.x)
            if not row_id or col_id != "#1":
                return
            vals = list(tree.item(row_id, "values"))
            if not vals:
                return
            acc_no = str(vals[1]).strip()
            if vals[0] == "[x]":
                vals[0] = "[ ]"
                checked_accounts.discard(acc_no)
            else:
                vals[0] = "[x]"
                checked_accounts.add(acc_no)
            tree.item(row_id, values=tuple(vals))

        def mark_selected():
            acc_no = selected_account()
            if not acc_no:
                messagebox.showwarning("No Selection", "Select an account in this window.", parent=win)
                return
            self.manager.mark_collected(acc_no)
            self.deposit_lists = []
            refresh_popup_table()
            self._refresh_all()

        def unmark_selected():
            acc_no = selected_account()
            if not acc_no:
                messagebox.showwarning("No Selection", "Select an account in this window.", parent=win)
                return
            self.manager.unmark_collected(acc_no)
            self.deposit_lists = []
            refresh_popup_table()
            self._refresh_all()

        def mark_all_visible():
            for row in tree.get_children():
                vals = tree.item(row, "values")
                if vals:
                    self.manager.mark_collected(str(vals[1]).strip())
            self.deposit_lists = []
            refresh_popup_table()
            self._refresh_all()

        def unmark_all_visible():
            for row in tree.get_children():
                vals = tree.item(row, "values")
                if vals:
                    self.manager.unmark_collected(str(vals[1]).strip())
            self.deposit_lists = []
            refresh_popup_table()
            self._refresh_all()

        def mark_checked():
            selected = _selected_checked_accounts()
            if not selected:
                messagebox.showwarning("No Selection", "Tick one or more checkboxes first.", parent=win)
                return
            for acc_no in selected:
                self.manager.mark_collected(acc_no)
            self.deposit_lists = []
            checked_accounts.clear()
            refresh_popup_table()
            self._refresh_all()

        def unmark_checked():
            selected = _selected_checked_accounts()
            if not selected:
                messagebox.showwarning("No Selection", "Tick one or more checkboxes first.", parent=win)
                return
            for acc_no in selected:
                self.manager.unmark_collected(acc_no)
            self.deposit_lists = []
            checked_accounts.clear()
            refresh_popup_table()
            self._refresh_all()

        actions = ttk.Frame(win, padding=8)
        actions.pack(fill=tk.X)
        ttk.Button(actions, text="Mark Checked Collected", command=mark_checked).pack(side=tk.LEFT, padx=4)
        ttk.Button(actions, text="Unmark Checked", command=unmark_checked).pack(side=tk.LEFT, padx=4)
        ttk.Button(actions, text="Close", command=win.destroy).pack(side=tk.RIGHT, padx=4)

        search_entry.bind("<KeyRelease>", lambda _e: refresh_popup_table())
        status_filter.bind("<<ComboboxSelected>>", lambda _e: refresh_popup_table())
        due_filter.bind("<<ComboboxSelected>>", lambda _e: refresh_popup_table())
        tree.bind("<Button-1>", toggle_row_checkbox)
        tree.bind("<Double-1>", lambda _e: mark_selected())
        refresh_popup_table()

    def _generate_deposit_lists(self):
        self._open_generate_list_window()

    def _build_adjusted_collected_accounts(self) -> List[RDAccount]:
        """Build collected accounts with denomination adjusted by per-account pay times."""
        adjusted_accounts: List[RDAccount] = []
        for acc in self.manager.collected_accounts:
            repeat_count = int(self.account_repeat_map.get(acc.account_no, 1))
            repeat_count = max(1, min(60, repeat_count))
            adjusted_accounts.append(
                replace(acc, denomination=acc.denomination * repeat_count)
            )
        return adjusted_accounts

    def _generate_deposit_lists_with_repeat(self, max_amt: float):
        adjusted_accounts = self._build_adjusted_collected_accounts()
        if not adjusted_accounts:
            self.deposit_lists = []
            self._refresh_preview()
            messagebox.showwarning(
                "No Lists",
                "No collected accounts found.\nMark depositors in 'Collected Window' first.",
            )
            return

        self.deposit_lists = DepositListGenerator.generate(adjusted_accounts, max_amt)
        self._refresh_preview()
        if not self.deposit_lists:
            messagebox.showwarning(
                "No Lists",
                "No collected accounts found.\nMark depositors in 'Collected Window' first.",
            )
        else:
            messagebox.showinfo(
                "Done",
                f"Generated {len(self.deposit_lists)} deposit list(s) from collected accounts.",
            )

    def _open_generate_list_window(self):
        win = tk.Toplevel(self.root)
        win.title("Generate Deposit Lists")
        win.geometry("430x170")
        win.transient(self.root)
        win.grab_set()

        frame = ttk.Frame(win, padding=12)
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            frame,
            text="Create lists from COLLECTED accounts (money with agent).",
        ).pack(anchor=tk.W, pady=(0, 10))

        max_var = tk.StringVar(value="20000")

        row1 = ttk.Frame(frame)
        row1.pack(fill=tk.X, pady=4)
        ttk.Label(row1, text="Max amount per list:", width=24).pack(side=tk.LEFT)
        ttk.Entry(row1, textvariable=max_var, width=14).pack(side=tk.LEFT)

        def run_generate():
            try:
                max_amt = float(max_var.get().strip())
                if max_amt <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror(
                    "Invalid Input",
                    "Enter a valid positive value for max amount.",
                    parent=win,
                )
                return

            self._generate_deposit_lists_with_repeat(max_amt)
            win.destroy()

        actions = ttk.Frame(frame)
        actions.pack(fill=tk.X, pady=(14, 0))
        ttk.Button(actions, text="Generate", command=run_generate).pack(side=tk.LEFT)
        ttk.Button(actions, text="Cancel", command=win.destroy).pack(side=tk.RIGHT)

    def _open_collected_accounts_window(self):
        win = tk.Toplevel(self.root)
        win.title("Collected Accounts")
        win.geometry("920x560")
        win.transient(self.root)
        win.grab_set()

        top = ttk.Frame(win, padding=8)
        top.pack(fill=tk.X)
        ttk.Label(top, text="Collected accounts only (not yet deposited):").pack(side=tk.LEFT)

        table_wrap = ttk.Frame(win, padding=(8, 0, 8, 8))
        table_wrap.pack(fill=tk.BOTH, expand=True)

        cols = ("account_no", "name", "amount", "months_paid", "priority", "due_date", "times")
        tree = ttk.Treeview(table_wrap, columns=cols, show="headings")
        tree.heading("account_no", text="Account No")
        tree.heading("name", text="Name")
        tree.heading("amount", text="Amount")
        tree.heading("months_paid", text="Months Paid")
        tree.heading("priority", text="Priority")
        tree.heading("due_date", text="Due Date")
        tree.heading("times", text="Pay Times")

        tree.column("account_no", width=120, anchor=tk.CENTER)
        tree.column("name", width=280, anchor=tk.W)
        tree.column("amount", width=120, anchor=tk.E)
        tree.column("months_paid", width=120, anchor=tk.CENTER)
        tree.column("priority", width=110, anchor=tk.CENTER)
        tree.column("due_date", width=120, anchor=tk.CENTER)
        tree.column("times", width=90, anchor=tk.CENTER)

        # --- Column sort state & logic ---
        ca_sort_state: Dict[str, bool] = {}

        ca_heading_labels = {
            "account_no": "Account No", "name": "Name", "amount": "Amount",
            "months_paid": "Months Paid", "priority": "Priority",
            "due_date": "Due Date", "times": "Pay Times",
        }

        def _ca_sort_key(col: str, value: str):
            val = str(value).strip()
            if col in ("amount",):
                cleaned = val.replace(",", "").replace("Cr.", "").replace("Cr", "").strip()
                try:
                    return float(cleaned)
                except ValueError:
                    return 0.0
            if col in ("months_paid", "times"):
                try:
                    return int(val)
                except ValueError:
                    return 0
            if col == "due_date":
                try:
                    return datetime.strptime(val, "%d-%b-%Y")
                except ValueError:
                    return datetime.min
            return val.lower()

        def _ca_sort_by_column(col: str):
            ascending = not ca_sort_state.get(col, False)
            ca_sort_state[col] = ascending
            rows = []
            for iid in tree.get_children():
                vals = tree.item(iid, "values")
                rows.append((iid, vals))
            col_idx = cols.index(col)
            rows.sort(
                key=lambda r: _ca_sort_key(col, r[1][col_idx] if col_idx < len(r[1]) else ""),
                reverse=not ascending,
            )
            for idx, (iid, _vals) in enumerate(rows):
                tree.move(iid, "", idx)
            arrow = " ▲" if ascending else " ▼"
            for c, label in ca_heading_labels.items():
                tree.heading(c, text=(label + arrow) if c == col else label)

        for col_name in ca_heading_labels:
            tree.heading(
                col_name,
                text=ca_heading_labels[col_name],
                command=lambda c=col_name: _ca_sort_by_column(c),
            )

        yscroll = ttk.Scrollbar(table_wrap, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)

        for acc in self.manager.collected_accounts:
            tree.insert(
                "",
                tk.END,
                values=(
                    acc.account_no,
                    acc.name,
                    f"{acc.denomination:,.0f}",
                    acc.months_paid,
                    acc.priority.value,
                    acc.due_date.strftime("%d-%b-%Y") if acc.due_date else "-",
                    self.account_repeat_map.get(acc.account_no, 1),
                ),
            )

        control_row = ttk.Frame(win, padding=(8, 0, 8, 4))
        control_row.pack(fill=tk.X)
        ttk.Label(control_row, text="Set pay times for selected account(s):").pack(side=tk.LEFT)
        times_var = tk.StringVar(value="1")
        times_combo = ttk.Combobox(
            control_row,
            textvariable=times_var,
            values=[str(i) for i in range(1, 61)],
            width=8,
            state="readonly",
        )
        times_combo.pack(side=tk.LEFT, padx=6)
        times_combo.current(0)

        def apply_times_to_selected():
            selected_rows = tree.selection()
            if not selected_rows:
                messagebox.showwarning("No Selection", "Select one or more accounts first.", parent=win)
                return
            times = int(times_var.get())
            for row_id in selected_rows:
                vals = list(tree.item(row_id, "values"))
                if not vals:
                    continue
                acc_no = str(vals[0]).strip()
                self.account_repeat_map[acc_no] = times
                vals[6] = times
                tree.item(row_id, values=tuple(vals))
            messagebox.showinfo("Updated", "Pay times updated for selected account(s).", parent=win)

        ttk.Button(control_row, text="Apply", command=apply_times_to_selected).pack(side=tk.LEFT)

        actions = ttk.Frame(win, padding=8)
        actions.pack(fill=tk.X)
        ttk.Button(actions, text="Generate List From This Window", command=self._open_generate_list_window).pack(side=tk.LEFT)
        ttk.Button(actions, text="Close", command=win.destroy).pack(side=tk.RIGHT)

    def _export_excel(self):
        if not self.deposit_lists:
            adjusted_accounts = self._build_adjusted_collected_accounts()
            self.deposit_lists = DepositListGenerator.generate(adjusted_accounts, 20000)
        if not self.deposit_lists:
            messagebox.showwarning(
                "No Data",
                "No collected accounts found.\nMark depositors in 'Collected Window' first.",
            )
            return

        default_name = f"RD_Deposit_Report_{date.today().strftime('%Y%m%d')}.xlsx"
        output = filedialog.asksaveasfilename(
            title="Save Excel report",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not output:
            return
        try:
            path = ExcelExporter.export_deposit_lists(self.deposit_lists, self.manager, output)
            base, _ext = os.path.splitext(output)
            txt_path = f"{base}_listwise_account_numbers.txt"

            lines: List[str] = []
            for idx, dlist in enumerate(self.deposit_lists, 1):
                acc_nos = [acc.account_no for acc in dlist.accounts]
                comma_separated = ",".join(acc_nos)
                lines.append(f"List {idx}: {comma_separated}")

            with open(txt_path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))

            messagebox.showinfo(
                "Export Complete",
                "Report exported to:\n"
                f"{os.path.abspath(path)}\n\n"
                "Account list saved to:\n"
                f"{os.path.abspath(txt_path)}",
            )
        except Exception as exc:
            messagebox.showerror("Export Failed", str(exc))

    def run(self):
        self.root.mainloop()


# ═══════════════════════════════════════════════════════════════
# SECTION 8: ENTRY POINT
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    if "--cli" in sys.argv:
        app = CLIInterface()
        app.run()
    else:
        app = GUIInterface()
        app.run()
